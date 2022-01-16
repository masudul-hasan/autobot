import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

"""
All the functions that is required to read a Excel Data Format. 

written by: jiaul_islam
"""

# Need to Change which worksheet need to select
SHEET_NAME = "Change_List"


#  Excel File Handler Class
class Read_Data:
    def __init__(self, file_path) -> None:
        """ Read the Excel Data as per Format """
        try:
            self._workbook: Workbook = load_workbook(read_only=True, filename=file_path, data_only=True)
            self._sheet: Worksheet = self._workbook.active
        except FileNotFoundError:
            print(
                f"File not found in {Path.cwd()}."
                f"\nMake sure you have 'data_driver' folder or files in the folder.\n"
                f"or for Draft NCR keep a draft.xlsx file in the data_driver folder.\n")
            sys.exit()
        except Exception as e:
            print(e)

    def change_sheet(self) -> None:
        """ Change the sheet """
        try:
            self._sheet = self._workbook.get_sheet_by_name(SHEET_NAME)
        except Exception as error:
            print(error)

    def get_number_change(self) -> int:
        """ get max SL no from the numbers """
        number_of_change = []
        for no in self._sheet.values:
            if no[0] is not None and no[0] != 'No':
                if no[0] != '0':
                    number_of_change.append(int(no[0]))
                else:
                    break
        return max(number_of_change)

    def parse_date(self, index: int) -> str:
        """ get the date from excel file """
        date = self._sheet['B' + str(index)]
        return date.value

    def parse_project_coordinator(self, index: int) -> str:
        """ get the project coordinator name from the excel """
        coordinator = self._sheet['C' + str(index)]
        return coordinator.value

    def parse_project_name(self, index: int) -> str:
        """ get the project name from excel file """
        project_name = self._sheet['D' + str(index)]
        return project_name.value

    def parse_change_activity(self, index: int):
        """ get the change activity from excel file """
        activity = self._sheet['E' + str(index)]
        return activity.value

    def parse_impact_list(self, index: int) -> str:
        """ get the impact list from excel file """
        impact_list = self._sheet['F' + str(index)]
        return impact_list.value

    def parse_service_type(self, index: int) -> str:
        """ get the service type from excel file """
        type_of_service = self._sheet['G' + str(index)]
        return type_of_service.value

    def parse_downtime_hour(self, index: int) -> str:
        """ get the downtime duration from excel file """
        hour_of_downtime = self._sheet['H' + str(index)]
        return hour_of_downtime.value

    @staticmethod
    def get_company_group() -> str:
        """ get the company group """
        return "e.co"

    @staticmethod
    def get_region() -> str:
        """ get the region from excel file """
        return "Asia"

    def parse_commercial_zone(self, index: int) -> str:
        """ get the commercial zone from excel file """
        commercial_zone = self._sheet['I' + str(index)]
        return commercial_zone.value

    def parse_change_manager(self, index: int) -> str:
        """ get the change manager from excel file """
        change_manager = self._sheet['K' + str(index)]
        return change_manager.value

    def get_last_empty_cell(self) -> int:
        """ Return the last empty row number """
        _last_blank_row = 0

        for row in range(1, self._sheet.max_row):
            if self._sheet[f"A{row}"].value is None:
                break
            else:
                _last_blank_row += 1
        return _last_blank_row

    def close_workbook(self) -> None:
        """ close the excel file """
        self._workbook.close()
